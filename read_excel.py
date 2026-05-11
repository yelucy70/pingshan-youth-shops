import openpyxl

wb = openpyxl.load_workbook('小店简介.xlsx')
ws = wb.active

print('Columns:', [cell.value for cell in ws[1]])
print()
for row in ws.iter_rows(min_row=2, max_row=31):
    values = [cell.value for cell in row]
    print(f"Row {row[0].row}: {values}")
